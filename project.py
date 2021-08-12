# -*- coding: utf-8 -*-

import datetime
import getpass
import zerosms
import openpyxl

from twilio.rest import Client



import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

#wb=Workbook()
#friends=wb.add_sheet("friends")
#relatives=wb.add_sheet("relatives")
#colleagues=wb.add_sheet("colleagues")
#wb.save("Python.xlsx")

def addFriends():
    name=str(input("Enter name of friend \t"))
    nickname=str(input("Enter "+name+"'s nickname \t"))
    contact=int(input("Enter "+name+"'s contact number \t"))
    email=str(input("Enter "+name+"'s email id \t"))
    bYear=int(input("Enter "+name+"'s year of birth \t"))
    bMonth=int(input("Enter "+name+"'s month of birth \t"))
    bDay=int(input("Enter "+name+"'s day of birth \t"))
    dob=datetime.date(bYear,bMonth,bDay)
    info=[name,contact,email,dob,dob.year,dob.month,dob.day,nickname]
    writeFriends(info)
    
def addRelatives():
    name=str(input("Enter name of relative \t"))
    contact=int(input("Enter "+name+"'s contact number \t"))
    email=str(input("Enter "+name+"'s email id \t"))
    bYear=int(input("Enter "+name+"'s year of birth \t"))
    bMonth=int(input("Enter "+name+"'s month of birth \t"))
    bDay=int(input("Enter "+name+"'s day of birth \t"))
    dob=datetime.date(bYear,bMonth,bDay)
    info=[name,contact,email,dob,dob.year,dob.month,dob.day]
    writeRelatives(info)

def addColleagues():
    name=str(input("Enter name of c \t"))
    contact=int(input("Enter "+name+"'s contact number \t"))
    email=str(input("Enter "+name+"'s email id \t"))
    bYear=int(input("Enter "+name+"'s year of birth \t"))
    bMonth=int(input("Enter "+name+"'s month of birth \t"))
    bDay=int(input("Enter "+name+"'s day of birth \t"))
    dob=datetime.date(bYear,bMonth,bDay)
    info=[name,contact,email,dob,dob.year,dob.month,dob.day]
    writeColleagues(info)

def writeFriends(info):
    wb=openpyxl.load_workbook("Python.xlsx")
    sheet=wb.get_sheet_by_name("friends")
    col=sheet.max_column
    for i in range(1,9):
        sheet.cell(row=i,column=col+1).value=info[0]
        del info[0]
    wb.save("Python.xlsx")
    print("/n Data saved successfully \n")


def writeRelatives(info):
    wb=openpyxl.load_workbook("Python.xlsx")
    sheet=wb.get_sheet_by_name("relatives")
    col=sheet.max_column
    for i in range(1,8):
        sheet.cell(row=i,column=col+1).value=info[0]
        del info[0]
    wb.save("Python.xlsx")
    print("/Data saved successfully \n")


def writeColleagues(info):
    wb=openpyxl.load_workbook("Python.xlsx")
    sheet=wb.get_sheet_by_name("colleagues")
    col=sheet.max_column
    for i in range(1,8):
        sheet.cell(row=i,column=col+1).value=info[0]
        del info[0]
    wb.save("Python.xlsx")
    print("\nData saved successfully \n")


def bdayChecker(sheetname):
    wb=openpyxl.load_workbook("Python.xlsx")
    sheet=wb.get_sheet_by_name(sheetname)
    bday=False
    for i in range(1,(sheet.max_column)+1):
        if sheet.cell(row=6,column=i).value==datetime.date.today().month and sheet.cell(row=7,column=i).value==datetime.date.today().day:
            eWishes(sheet.cell(row=1,column=i).value,sheet.cell(row=3,column=i).value)
            mobileWishes(sheet.cell(row=1,column=i).value,sheet.cell(row=2,column=i).value)
            bday=True
    if bday==False:
        print("\n No Birthdays Today in "+sheetname+" list! \n")
    return True

def eWishes(name,email):
    sender="smd065496@gmail.com"
    receiver=email
    subject="Birthday Wishes"
    
    msg=MIMEMultipart()
    msg["From"]=sender
    msg["To"]=receiver
    msg["Subject"]=subject
    
    body=str(input("Its "+name+"'s Birthday ! Enter a message to wish them \n"))
    msg.attach(MIMEText(body,"plain"))
    
    choice=str(input("Want to attach any files? If yes enter 'y' \t"))
    if choice=="y":
        filename=str(input("Enter filename with path address \n"))
        attachment=open(filename,"rb")
        
        part=MIMEBase("application","octet-stream")
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition","attachment; filename "+filename)
        
        msg.attach(part)
        text=msg.as_string()
        server=smtplib.SMTP("smtp.gmail.com",587)
        password=str(input("Enter password \n"))
        server.starttls()
        server.login(sender,password)
        
        server.sendmail(sender,receiver,text)
        server.quit()
    else:
        text=msg.as_string()
        server=smtplib.SMTP("smtp.gmail.com",587)
        password=getpass.getpass("Enter password for email-id "+sender+"\n")
        server.starttls()
        server.login(sender,password)
        
        server.sendmail(sender,receiver,text)
        server.quit()

def mobileWishes(name,mobile):
    msg=str(input(" Enter a message to wish them on mobile \n"))
    
    account_sid = 'ACf0472249e8ae60f130d724f5eaaf9f8c'#enter your auth sid of twilio account
    auth_token = '6b5c1932d98ca2aac1a609b896e89ecb' #enter your auth token of twilio account
    client = Client(account_sid, auth_token)
    

    message = client.messages         .create(
             body=msg,
             from_= +15036837169 , #enter your twilio number
             to= mobile
         )

    print("Birthday Message has sent successfully...")


            

choice='1'
while choice=='1'or choice=='2'or choice=='3':
    bdayChecker("friends")
    bdayChecker("relatives")
    bdayChecker("colleagues")
    print("\nEnter your choice to add info about")
    choice=str(input("1.Friend\n2.Relative\n3.Colleague\n"))
    if choice=='1':
        addFriends()
    elif choice=='2':
        addRelatives()
    elif choice=='3':
        addColleagues()
    else:
        break
